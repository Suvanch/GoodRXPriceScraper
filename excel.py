from openpyxl import load_workbook

myFileName=r'Order_List.xlsx'
#load the workbook, and put the sheet into a variable
wb = load_workbook(filename=myFileName)
ws = wb['Sheet1']

#max_row is a sheet function that gets the last row in a sheet.
newRowLocation = ws.max_row +1

#write to the cell you want, specifying row and column, and value :-)
ws.cell(column=1,row=newRowLocation, value="aha! a new entry at the end")
wb.save(filename=myFileName)
wb.close()